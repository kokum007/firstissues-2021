# importing openpyxl module
import openpyxl as xl;

# opening the source excel file
filename ="INTERVIEW SELECTION  TEST.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[1]
ws2 = wb1.worksheets[2]



# copying the cell values from source
# excel file to destination excel file

#component values
for i in range (8, 17):
	
		# reading cell value from source excel file
		c = ws1.cell(row = i, column = 1)

		# writing the read value to destination excel file
		ws2.cell(row = i-6, column = 5).value = c.value

#component decriptions        
for i in range (8, 17):
	
		# reading cell value from source excel file
		c = ws1.cell(row = i, column = 2)

		# writing the read value to destination excel file
		ws2.cell(row = i-6, column = 6).value = c.value


#cost price per kg
for i in range (8, 17):
	
		# reading cell value from source excel file
		c = ws1.cell(row = i, column = 3)

		# writing the read value to destination excel file
		ws2.cell(row = i-6, column = 9).value = c.value        

#version
for i in range (8, 17):
	
		# reading cell value from source excel file
		c = ws1.cell(row = 1, column = 2)
        
		# writing the read value to destination excel file
		ws2.cell(row = i-6, column = 1).value = c.value


#period
for i in range (8, 17):
	
		# reading cell value from source excel file
		c = ws1.cell(row = 2, column = 2)
        
		# writing the read value to destination excel file
		ws2.cell(row = i-6, column = 8).value = c.value

#profit center
for i in range (8, 17):
	
		# reading cell value from source excel file
		c = ws1.cell(row = 3, column = 2)
        
		# writing the read value to destination excel file
		ws2.cell(row = i-6, column = 2).value = c.value


#profit center decription
for i in range (8, 17):
	
		# reading cell value from source excel file
		c = ws1.cell(row = 4, column = 2)
        
		# writing the read value to destination excel file
		ws2.cell(row = i-6, column = 3).value = c.value


#customer id
for i in range (8, 17):
	
		# reading cell value from source excel file
		c = i-7
        # writing the read value to destination excel file
		ws2.cell(row = i-6, column = 4).value = c



# saving the destination excel file
wb1.save(str(filename))
